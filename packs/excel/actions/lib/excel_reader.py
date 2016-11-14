"""
Copyright 2016 Brocade Communications Systems, Inc.
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at
    http://www.apache.org/licenses/LICENSE-2.0
Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

from openpyxl import Workbook, load_workbook
import os
import time


class ExcelReader(object):
    def __init__(self, excel_file, lock=False, lock_retries=3, lock_delay=1):
        ''' Loads the excel configuration file '''
        self._filename = excel_file
        self._lock = lock

        # Set class variables
        self._strict = True
        self._wb = None
        self._ws = None
        self._sheet_name = None
        self._key_column = -1
        self._keys = []
        self._var_name_row = -1
        self._variable_start_column = -1
        self._variable_end_column = -1
        self._data_start_row = -1
        self._data_end_row = -1

        if lock:
            # is file already locked?
            attempts = 0
            while os.path.isfile("%s.lock" % excel_file):
                time.sleep(lock_delay)
                attempts += 1
                if attempts >= lock_retries:
                    raise IOError("File is locked")
            # it's now ours, so lock it
            try:
                open("%s.lock" % excel_file, 'a').close()
            except IOError:
                raise IOError("Could not lock spreadsheet '%s'" % excel_file)
        try:
            self._wb = load_workbook(excel_file, data_only=True)
        except IOError:
            # Create a blank workbook
            self._wb = Workbook()

    def __del__(self):
        ''' Get rid of lock file if as cleanup '''
        try:
            self._unlock_file()
        except OSError:
            pass

    def _unlock_file(self):
        ''' Attempts to remove the locked file '''
        if self._lock:
            os.remove("%s.lock" % self._filename)
            self._lock = False

    def save(self):
        ''' Saves spreadsheet to disk '''
        if not self._lock:
            raise IOError("Trying to save file not locked for modification")
        self._wb.save(self._filename)
        self._unlock_file()

    def get_sheets(self):
        ''' Returns an array of the sheet names '''
        sheets = [sheet.title for sheet in self._wb]
        return sheets

    def get_keys(self):
        ''' Returns the keys in the sheet '''
        if not self._ws:
            raise KeyError("Sheet not specified")
        keys = list(self._keys)
        return keys

    def get_variable_names(self):
        ''' Returns the variable names in the sheet '''
        if not self._ws:
            raise KeyError("Sheet not specified")
        variable_names = []
        for col in range(self._variable_start_column,
                         self._variable_end_column):

            variable_names.append(self._ws.cell(column=col,
                                  row=self._var_name_row).value)

        return variable_names

    def _set_key_column(self, key_column):
        ''' Set the key column '''
        self._key_column = key_column
        self._keys = {}
        row = self._data_start_row
        while True:
            key = self._ws.cell(column=self._key_column, row=row)
            if key.value:
                # Check for duplicate key
                if key.value not in self._keys:
                    self._keys[key.value] = row
                else:
                    self._unlock_file()
                    raise ValueError("Duplicate key '%s' found at row '%s'." %
                                     (key.value, row))
            else:
                self._data_end_row = row
                break
            row += 1

    def _set_variable_start_column(self, variable_start_column):
        ''' Set the template column in the template '''
        col = self._variable_start_column = variable_start_column
        while True:
            cell = self._ws.cell(column=col, row=self._var_name_row)
            if cell.value:
                col += 1
            else:
                self._variable_end_column = col
                break

    def set_sheet(self, sheet_name, key_column=1, var_name_row=1,
                  strict=False):
        ''' Sets the current sheet '''
        self._sheet_name = sheet_name
        self._strict = strict
        try:
            self._ws = self._wb.get_sheet_by_name(self._sheet_name)
        except KeyError:
            if not self._strict:
                if self._lock:
                    self._ws = self._wb.create_sheet(self._sheet_name)
                else:
                    raise IOError("File not locked for modification")
            else:
                self._unlock_file()
                raise KeyError("Sheet '%s' not found" % self._sheet_name)

        self._var_name_row = var_name_row
        self._data_start_row = var_name_row + 1

        self._key_column = key_column
        self._set_key_column(key_column)

        self._variable_start_column = key_column + 1
        self._variable_end_column = self._variable_start_column
        self._set_variable_start_column(key_column + 1)

    def get_row_for_key(self, key):
        ''' Returns the row for a given key, or -1 if not matched '''
        if len(key) > 255:
            self._unlock_file()
            raise ValueError("Key exceeds 255 characters")
        if key in self._keys:
            return self._keys[key]
        return -1

    def get_variables_for_key(self, key):
        ''' Returns the variables for a particular key '''
        if not self._ws:
            raise NameError("Sheet not specified")
        variables = {}
        col = self._variable_start_column
        row = self.get_row_for_key(key)
        while row >= 0 and col <= self._variable_end_column:
            variable = self._ws.cell(column=col, row=self._var_name_row)
            if variable.value:
                variables[variable.value] = self._ws.cell(column=col,
                                                          row=row).value
            col += 1
        return variables

    def get_last_row(self):
        ''' Returns the last row of data '''
        return self._data_end_row

    def set_values_for_variables(self, key, dictionary):
        ''' Sets values for variables for a given key '''
        # Make sure the sheet has been specified
        if not self._ws:
            raise NameError("Sheet not specified")
        # Check to make sure file was locked before allowing modifications
        if not self._lock:
            raise IOError("File not locked for modification")

        # Build a dictionary of variables in the spreadsheet
        variables = {}
        col = self._variable_start_column
        while True:
            variable = self._ws.cell(column=col, row=self._var_name_row)
            if variable.value:
                variables[variable.value] = col
            else:
                self._variable_end_column = col
                break
            col += 1

        # Find row for key
        row = self.get_row_for_key(key)
        if row == -1:
            row = self._data_end_row
            self._ws.cell(column=self._key_column, row=row).value = key
            self._data_end_row += 1
            self._keys[key] = row

        # Fill in values
        for k, value in dictionary.items():
            # is key or value greater the excel 256 character limit?
            if len(k) > 255:
                self._unlock_file()
                raise ValueError("Variable name exceeds 255 characters")
            if len(k) == 0:
                self._unlock_file()
                raise ValueError("Variable name is blank")
            if len(value) > 32767:
                self._unlock_file()
                raise ValueError("Variable value exceeds 32,767 characters")
            if not k == key:
                if k in variables:
                    self._ws.cell(column=variables[k], row=row).value = value
                else:
                    if not self._strict:
                        # Add a new column at the end with variable
                        self._ws.cell(column=self._variable_end_column,
                                      row=self._var_name_row).value = k
                        self._ws.cell(column=self._variable_end_column,
                                      row=row).value = value
                        self._variable_end_column += 1
                        if self._variable_end_column > 16384:
                            raise ValueError("Exceeded max # of columns")
                    else:
                        self._unlock_file()
                        raise ValueError("Column '%s' missing in sheet '%s'" %
                                         (k, self._sheet_name))

if __name__ == "__main__":
    pass
